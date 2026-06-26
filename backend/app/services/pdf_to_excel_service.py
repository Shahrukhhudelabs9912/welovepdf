"""
PDF to Excel conversion service for WeLovePDF.
Uses pdfplumber for table extraction and pandas/openpyxl for Excel generation.
"""
import io
import os
import re
import uuid
import logging
import tempfile
from typing import List, Optional, Tuple

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class PdfToExcelService:
    """Service to convert PDF files containing tables to Excel (.xlsx) format."""

    # Maximum number of pages to process (safety limit)
    MAX_PAGES = 200

    # pdfplumber table-detection strategies, tried in order.
    # 1) lines: works on PDFs with visible borders.
    # 2) text:  works on whitespace-aligned tables (bank statements, invoices, etc).
    TABLE_STRATEGIES = [
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
        {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "snap_tolerance": 4,
            "join_tolerance": 4,
            "edge_min_length": 3,
            "min_words_vertical": 2,
            "min_words_horizontal": 1,
        },
    ]

    @staticmethod
    def _clean_table(raw_table) -> Optional[List[List[str]]]:
        """Normalize a raw pdfplumber table to a list of equal-width rows."""
        if not raw_table:
            return None
        cleaned = []
        for row in raw_table:
            cleaned_row = [
                (str(cell).strip() if cell is not None else "") for cell in row
            ]
            if any(cell != "" for cell in cleaned_row):
                cleaned.append(cleaned_row)
        if not cleaned:
            return None
        max_cols = max(len(r) for r in cleaned)
        # Drop tables that are really just a single column of stray text.
        if max_cols < 2:
            return None
        normalized = [r + [""] * (max_cols - len(r)) for r in cleaned]
        return normalized

    @staticmethod
    def _text_to_rows(text: str) -> List[List[str]]:
        """Fallback: split page text into a single-column row set."""
        rows: List[List[str]] = []
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            # Break on 2+ spaces — typical column separator in PDF text dumps.
            parts = [p.strip() for p in re.split(r"\s{2,}|\t+", stripped) if p.strip()]
            if not parts:
                parts = [stripped]
            rows.append(parts)
        if not rows:
            return rows
        max_cols = max(len(r) for r in rows)
        return [r + [""] * (max_cols - len(r)) for r in rows]

    @staticmethod
    def _dedupe_headers(header: List[str]) -> List[str]:
        """Make header names unique and non-empty. Duplicate or empty headers
        confuse pandas (df.name, df[col] return Series unexpectedly)."""
        seen: dict = {}
        result: List[str] = []
        for i, raw in enumerate(header):
            name = (str(raw).strip() if raw is not None else "") or f"Col{i + 1}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 1
            result.append(name)
        return result

    @staticmethod
    def extract_tables_from_pdf(pdf_bytes: bytes) -> List[Tuple[str, pd.DataFrame]]:
        """
        Extract tables from a PDF file using pdfplumber.

        Tries border-based detection first, then whitespace-based detection,
        and finally falls back to a plain text grid so the user always gets
        an Excel file as long as the PDF has selectable text.

        Returns:
            List of (sheet_name, DataFrame) tuples.

        Raises:
            ValueError: Only when the PDF has no extractable text at all
                (likely a scanned/image-only PDF).
        """
        tables: List[Tuple[str, pd.DataFrame]] = []
        any_text_found = False

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)
            logger.info(f"PDF has {total_pages} pages. Extracting tables...")

            if total_pages > PdfToExcelService.MAX_PAGES:
                logger.warning(
                    f"PDF has {total_pages} pages, limiting to {PdfToExcelService.MAX_PAGES}"
                )

            for page_num, page in enumerate(pdf.pages[:PdfToExcelService.MAX_PAGES], start=1):
                page_tables = []
                for strategy in PdfToExcelService.TABLE_STRATEGIES:
                    try:
                        found = page.extract_tables(table_settings=strategy)
                    except Exception as e:
                        logger.debug(
                            f"Page {page_num}: strategy {strategy} failed: {e}"
                        )
                        found = []
                    if found:
                        page_tables = found
                        logger.info(
                            f"Page {page_num}: Found {len(found)} table(s) using "
                            f"{strategy.get('vertical_strategy')}/"
                            f"{strategy.get('horizontal_strategy')} strategy"
                        )
                        break

                page_text = page.extract_text() or ""
                if page_text.strip():
                    any_text_found = True

                if page_tables:
                    for table_idx, raw in enumerate(page_tables):
                        normalized = PdfToExcelService._clean_table(raw)
                        if not normalized:
                            continue
                        header = PdfToExcelService._dedupe_headers(normalized[0])
                        data_rows = normalized[1:] if len(normalized) > 1 else []
                        df = pd.DataFrame(data_rows, columns=header)
                        sheet_name = f"Page{page_num}_Table{table_idx + 1}"
                        tables.append((sheet_name, df))
                    continue

                # No table on this page — fall back to text grid so the user
                # still gets the page content in their spreadsheet.
                if page_text.strip():
                    text_rows = PdfToExcelService._text_to_rows(page_text)
                    if text_rows:
                        max_cols = len(text_rows[0])
                        columns = [f"Col{i + 1}" for i in range(max_cols)]
                        df = pd.DataFrame(text_rows, columns=columns)
                        sheet_name = f"Page{page_num}_Text"
                        tables.append((sheet_name, df))
                        logger.info(
                            f"Page {page_num}: No table detected, used text fallback "
                            f"({len(text_rows)} row(s))"
                        )

        if not tables:
            if not any_text_found:
                raise ValueError(
                    "This PDF appears to contain only images (scanned document) and has "
                    "no selectable text. Please run it through the 'Fix Scanned PDF' (OCR) "
                    "tool first, then try PDF-to-Excel again."
                )
            raise ValueError(
                "Unable to extract any data from this PDF. The PDF may be empty or "
                "use an unsupported layout."
            )

        logger.info(f"Total sheets extracted: {len(tables)}")
        return tables

    @staticmethod
    def convert_to_excel(
        pdf_bytes: bytes,
        original_filename: str = "document.pdf",
    ) -> tuple[bytes, str]:
        """
        Convert a PDF file to an Excel workbook.

        Args:
            pdf_bytes: Raw PDF file content.
            original_filename: Original PDF filename for deriving output name.

        Returns:
            Tuple of (excel_bytes, output_filename).

        Raises:
            ValueError: If conversion fails (e.g., no tables found).
            Exception: For unexpected errors.
        """
        logger.info(f"Starting PDF to Excel conversion for: {original_filename}")

        try:
            # Step 1: Extract tables
            tables = PdfToExcelService.extract_tables_from_pdf(pdf_bytes)

            # Step 2: Build Excel workbook
            wb = Workbook()
            # Remove default sheet
            default_sheet = wb.active
            if default_sheet:
                wb.remove(default_sheet)

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            cell_font = Font(name="Calibri", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)

            sheet_names_used: set = set()

            for idx, (raw_sheet_name, df) in enumerate(tables):
                # Sanitize sheet name (Excel: max 31 chars, no special chars)
                sanitized = "".join(
                    c for c in str(raw_sheet_name) if c.isalnum() or c in (" ", "_", "-")
                )[:31]
                base_name = sanitized or f"Table_{idx + 1}"

                sheet_name = base_name
                counter = 1
                while sheet_name in sheet_names_used:
                    sheet_name = f"{base_name[:28]}_{counter}"
                    counter += 1
                sheet_names_used.add(sheet_name)

                ws = wb.create_sheet(title=sheet_name)

                # Write DataFrame to worksheet using openpyxl for formatting
                for r_idx, row in enumerate(
                    dataframe_to_rows(df, index=False, header=True), start=1
                ):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = thin_border

                        if r_idx == 1:  # Header row
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        else:
                            cell.font = cell_font
                            cell.alignment = cell_alignment

                # Auto-fit column widths (approximate)
                for col_cells in ws.columns:
                    max_length = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    # Cap at 50 characters to avoid overly wide columns
                    adjusted_width = min(max_length + 4, 50)
                    ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

                # Freeze header row
                ws.freeze_panes = "A2"

            # If only one table, name the file nicely
            base_filename = os.path.splitext(original_filename)[0]
            output_filename = f"{base_filename}_converted.xlsx"

            # Save workbook to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()

            logger.info(
                f"PDF to Excel conversion complete: {len(tables)} table(s), "
                f"{len(excel_bytes)} bytes output"
            )
            return excel_bytes, output_filename

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"PDF to Excel conversion failed: {str(e)}", exc_info=True)
            raise Exception(f"Conversion failed: {str(e)}")