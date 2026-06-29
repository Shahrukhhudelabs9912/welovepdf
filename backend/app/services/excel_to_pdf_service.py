"""
Excel to PDF conversion service for PDFOrca.

Primary path: render the *actual* spreadsheet to PDF with LibreOffice headless
(the same engine word-to-pdf / powerpoint-to-pdf already use). This preserves
the real Excel layout — fonts, colours, merged cells, column widths, number
formats — and matches what iLovePDF / SmallPDF produce. Before handing the
file to LibreOffice we force every sheet to "fit all columns on one page" so
wide, data-heavy sheets scale down to the page instead of being cut off.

Fallback path: if LibreOffice is unavailable, rebuild a basic table with
pandas + reportlab (lower fidelity, but better than failing).
"""
import io
import uuid
import logging
import os
import subprocess
import tempfile
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    KeepTogether,
)
from reportlab.platypus.flowables import HRFlowable

logger = logging.getLogger(__name__)

# LibreOffice binary is resolved at runtime via resolve_libreoffice_path()
# (env LIBREOFFICE_PATH → PATH → OS default), shared with Word/PowerPoint.
from app.utils.concurrency import resolve_libreoffice_path


class ExcelToPdfService:
    """Service to convert Excel (.xlsx/.xls) files to PDF format."""

    MAX_ROWS_PER_PAGE = 45  # Approximate rows per PDF page
    MAX_COLUMNS = 30  # Safety limit for columns

    @staticmethod
    def read_excel_sheets(excel_bytes: bytes, original_filename: str = "spreadsheet.xlsx") -> dict:
        """
        Read all sheets from an Excel file with automatic engine detection.

        Args:
            excel_bytes: Raw Excel file content.
            original_filename: Original filename for engine detection.

        Returns:
            Dictionary mapping sheet name to pandas DataFrame.

        Raises:
            ValueError: If the file is empty or unreadable.
        """
        sheets_data: dict = {}
        filename_lower = original_filename.lower()

        # Determine engine based on file extension
        if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
            # .xls files (legacy Excel 97-2004 format)
            engines_to_try = ['xlrd', 'openpyxl', None]
        elif filename_lower.endswith('.xlsm'):
            engines_to_try = ['openpyxl', None]
        elif filename_lower.endswith('.xlsx'):
            engines_to_try = ['openpyxl', 'xlrd', None]
        else:
            # Unknown extension: try all engines plus auto-detect for CSV/TSV
            engines_to_try = [None, 'openpyxl', 'xlrd']

        last_error = None

        for engine in engines_to_try:
            try:
                logger.info(f"Trying to read Excel file with engine: {engine or 'auto'}")
                xl = pd.ExcelFile(io.BytesIO(excel_bytes), engine=engine)
                sheet_names = xl.sheet_names
                logger.info(f"Excel file has {len(sheet_names)} sheet(s): {sheet_names}")

                for sheet_name in sheet_names:
                    df = pd.read_excel(xl, sheet_name=sheet_name)

                    # Drop fully empty rows and columns
                    df = df.dropna(how="all").dropna(axis=1, how="all")

                    if df.empty:
                        logger.warning(f"Sheet '{sheet_name}' is empty after cleaning, skipping.")
                        continue

                    # Limit columns for rendering
                    if len(df.columns) > ExcelToPdfService.MAX_COLUMNS:
                        logger.warning(
                            f"Sheet '{sheet_name}' has {len(df.columns)} columns, "
                            f"limiting to {ExcelToPdfService.MAX_COLUMNS}"
                        )
                        df = df.iloc[:, :ExcelToPdfService.MAX_COLUMNS]

                    # Convert all values to strings for PDF rendering
                    df = df.fillna("").astype(str)

                    sheets_data[sheet_name] = df

                if not sheets_data:
                    raise ValueError(
                        "No readable data found in the Excel file. "
                        "The file may be empty or contain only empty sheets."
                    )

                return sheets_data

            except ValueError:
                raise
            except ImportError as e:
                last_error = e
                logger.warning(f"Engine '{engine}' not available: {e}. Trying next engine...")
                continue
            except Exception as e:
                last_error = e
                logger.warning(f"Engine '{engine}' failed: {e}. Trying next engine...")
                continue

        # If we get here, try CSV/TSV as a last resort for mislabeled files
        try:
            logger.info("Attempting to read as CSV/TSV (file may be mislabeled)")
            try:
                df = pd.read_csv(io.BytesIO(excel_bytes), sep=None, engine='python')
            except Exception:
                df = pd.read_csv(io.BytesIO(excel_bytes))

            df = df.dropna(how="all").dropna(axis=1, how="all")
            if not df.empty:
                df = df.fillna("").astype(str)
                sheets_data["Sheet1"] = df
                logger.info("Successfully read file as CSV/TSV")
                return sheets_data
        except Exception:
            pass

        raise ValueError(
            f"Could not read Excel file. All engines failed. Last error: {last_error}. "
            "Supported formats: .xlsx, .xls, .xlsm, .csv, .tsv"
        )

    @staticmethod
    def _clean_cell_value(value) -> str:
        """Clean and truncate cell values for PDF rendering."""
        s = str(value).strip()
        # Truncate very long cell values
        if len(s) > 200:
            s = s[:197] + "..."
        # Replace newlines with spaces for table rendering
        s = s.replace("\n", " ").replace("\r", " ")
        return s

    @staticmethod
    def convert_to_pdf(
        excel_bytes: bytes,
        original_filename: str = "document.xlsx",
    ) -> tuple[bytes, str]:
        """
        Convert an Excel file to a PDF document.

        Tries LibreOffice headless first (high-fidelity, matches the real
        spreadsheet layout like iLovePDF / SmallPDF). Falls back to the
        reportlab table renderer if LibreOffice isn't available or fails.

        Returns:
            Tuple of (pdf_bytes, output_filename).
        """
        logger.info(f"Starting Excel to PDF conversion for: {original_filename}")
        base_filename = os.path.splitext(original_filename)[0]
        output_filename = f"{base_filename}.pdf"

        # Primary: LibreOffice headless — same engine as word/powerpoint-to-pdf.
        if resolve_libreoffice_path():
            try:
                pdf_bytes = ExcelToPdfService._convert_with_libreoffice(
                    excel_bytes, original_filename
                )
                logger.info(
                    f"Excel to PDF (LibreOffice) complete: {len(pdf_bytes)} bytes"
                )
                return pdf_bytes, output_filename
            except Exception as e:
                logger.warning(
                    f"LibreOffice Excel-to-PDF failed ({e}); "
                    "falling back to reportlab renderer."
                )
        else:
            logger.warning(
                "LibreOffice not found; using reportlab fallback for Excel-to-PDF."
            )

        # Fallback: rebuild a basic table with pandas + reportlab.
        return ExcelToPdfService._convert_with_reportlab(excel_bytes, original_filename)

    @staticmethod
    def _prepare_workbook_for_print(input_path: Path) -> None:
        """
        Force every sheet to scale all columns onto a single page width (and
        landscape for wide sheets). This is what stops large, many-column
        spreadsheets from being clipped at the right edge — LibreOffice honours
        these print settings when it exports to PDF.

        Best-effort: skipped silently for non-.xlsx inputs.
        """
        if input_path.suffix.lower() not in (".xlsx", ".xlsm"):
            return
        try:
            wb = load_workbook(input_path)
        except Exception as e:
            logger.info(f"Could not open workbook to set print scaling: {e}")
            return

        try:
            from openpyxl.worksheet.properties import PageSetupProperties
        except Exception:
            PageSetupProperties = None

        # Up to this many columns: safe to fit on one portrait page.
        NARROW_FIT_COLS = 6
        # Up to this many columns: still fit on one page, but in landscape.
        LANDSCAPE_FIT_COLS = 12
        # Mild zoom-out for wide sheets — keeps more columns per page while
        # staying legible (LibreOffice paginates the rest onto extra pages).
        WIDE_SHEET_SCALE = 75  # percent

        for ws in wb.worksheets:
            try:
                max_col = ws.max_column or 1

                ws.page_margins.left = 0.3
                ws.page_margins.right = 0.3
                ws.page_margins.top = 0.5
                ws.page_margins.bottom = 0.5

                if max_col <= NARROW_FIT_COLS:
                    # Narrow sheet: one page wide, portrait, font stays large.
                    ws.page_setup.orientation = "portrait"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                    if PageSetupProperties is not None:
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                elif max_col <= LANDSCAPE_FIT_COLS:
                    # Medium sheet: one page wide but landscape for legibility.
                    ws.page_setup.orientation = "landscape"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                    if PageSetupProperties is not None:
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                else:
                    # Wide sheet: do NOT crush everything onto one page (that is
                    # what made the font unreadable). Keep cells legible and let
                    # the extra columns flow onto additional pages.
                    ws.page_setup.orientation = "landscape"
                    ws.page_setup.fitToWidth = 0
                    ws.page_setup.fitToHeight = 0
                    if PageSetupProperties is not None:
                        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
                    ws.page_setup.scale = WIDE_SHEET_SCALE

                # Repeat the header row on every page so multi-page output stays usable.
                if ws.max_row and ws.max_row > 1:
                    ws.print_title_rows = "1:1"
            except Exception as e:
                logger.info(f"Print-setup tweak skipped for a sheet: {e}")

        wb.save(input_path)

    @staticmethod
    def _convert_with_libreoffice(excel_bytes: bytes, original_filename: str) -> bytes:
        """Render the real spreadsheet to PDF via LibreOffice headless."""
        safe_name = os.path.basename(original_filename) or "spreadsheet.xlsx"
        if not os.path.splitext(safe_name)[1]:
            safe_name += ".xlsx"

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / safe_name
            input_path.write_bytes(excel_bytes)

            # Force fit-to-width / landscape before exporting so data-heavy
            # sheets scale to the page instead of being cut off.
            ExcelToPdfService._prepare_workbook_for_print(input_path)

            libreoffice_path = resolve_libreoffice_path()
            if not libreoffice_path:
                raise RuntimeError("LibreOffice is not installed on the server.")

            # Unique profile dir so parallel LibreOffice runs don't collide.
            profile_dir = tmp_path / f"lo_profile_{uuid.uuid4().hex[:8]}"
            cmd = [
                libreoffice_path,
                "--headless",
                f"-env:UserInstallation=file:///{profile_dir.as_posix().lstrip('/')}",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(tmp_path),
                str(input_path),
            ]
            try:
                result = subprocess.run(
                    cmd, capture_output=True, text=True, timeout=180
                )
            except subprocess.TimeoutExpired:
                raise RuntimeError("Excel to PDF conversion timed out.")

            if result.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice failed: {result.stderr[:200] or 'unknown error'}"
                )

            pdfs = list(tmp_path.glob("*.pdf"))
            if not pdfs:
                raise RuntimeError("Conversion completed but no PDF was generated.")
            return pdfs[0].read_bytes()

    @staticmethod
    def _convert_with_reportlab(
        excel_bytes: bytes,
        original_filename: str = "document.xlsx",
    ) -> tuple[bytes, str]:
        """
        Lower-fidelity fallback: read the data with pandas and rebuild a
        styled table with reportlab. Used only when LibreOffice is missing
        or errors out.

        Raises:
            ValueError: If conversion fails.
        """
        logger.info(f"Starting Excel to PDF (reportlab fallback) for: {original_filename}")

        try:
            # Step 1: Read Excel sheets with engine auto-detection
            sheets_data = ExcelToPdfService.read_excel_sheets(excel_bytes, original_filename)

            # Step 2: Build PDF
            pdf_buffer = io.BytesIO()

            # Determine page orientation based on column count
            max_cols = max(len(df.columns) for df in sheets_data.values())
            page_size = landscape(A4) if max_cols > 6 else A4

            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=page_size,
                topMargin=0.5 * inch,
                bottomMargin=0.5 * inch,
                leftMargin=0.4 * inch,
                rightMargin=0.4 * inch,
                title=f"Excel to PDF: {original_filename}",
                author="PDFOrca",
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.HexColor("#1a1a2e"),
            )
            sheet_title_style = ParagraphStyle(
                "SheetTitle",
                parent=styles["Heading2"],
                fontSize=12,
                spaceAfter=6,
                textColor=colors.HexColor("#4472C4"),
            )

            story = []

            # Main title
            base_filename = os.path.splitext(original_filename)[0]
            story.append(Paragraph(f"Excel to PDF: {base_filename}", title_style))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CCCCCC")))
            story.append(Spacer(1, 0.2 * inch))

            # Available width for tables
            avail_width = page_size[0] - doc.leftMargin - doc.rightMargin

            sheet_names = list(sheets_data.keys())

            for sheet_idx, (sheet_name, df) in enumerate(sheets_data.items()):
                if sheet_idx > 0:
                    story.append(PageBreak())

                # Sheet title
                story.append(Paragraph(f"Sheet: {sheet_name}", sheet_title_style))
                story.append(Spacer(1, 0.1 * inch))

                # Build table data
                table_data = []
                # Header row
                header_row = [Paragraph(f"<b>{ExcelToPdfService._clean_cell_value(col)}</b>", styles["Normal"]) for col in df.columns]
                table_data.append(header_row)

                # Data rows
                for _, row in df.iterrows():
                    data_row = [
                        Paragraph(ExcelToPdfService._clean_cell_value(val), styles["Normal"])
                        for val in row.values
                    ]
                    table_data.append(data_row)

                # Calculate column widths
                num_cols = len(df.columns)
                if num_cols > 0:
                    col_width = avail_width / num_cols
                    # Ensure minimum column width
                    col_width = max(col_width, 0.8 * inch)
                else:
                    col_width = avail_width

                col_widths = [col_width] * num_cols

                # Create table
                tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

                # Table styling
                style_commands = [
                    # Header styling
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    # Grid
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
                    ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#2F5496")),
                    # Alignment
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    # Padding
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]

                # Alternating row colors
                for i in range(1, len(table_data)):
                    if i % 2 == 0:
                        style_commands.append(
                            ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F2F7FC"))
                        )

                tbl.setStyle(TableStyle(style_commands))
                story.append(tbl)

                # Sheet info
                story.append(Spacer(1, 0.15 * inch))
                story.append(
                    Paragraph(
                        f"<i>Sheet '{sheet_name}': {len(df)} rows × {len(df.columns)} columns</i>",
                        styles["Italic"],
                    )
                )

            # Build the PDF
            doc.build(story)
            pdf_buffer.seek(0)
            pdf_bytes = pdf_buffer.getvalue()

            output_filename = f"{base_filename}_converted.pdf"

            logger.info(
                f"Excel to PDF conversion complete: {len(sheet_names)} sheet(s), "
                f"{len(pdf_bytes)} bytes output"
            )
            return pdf_bytes, output_filename

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Excel to PDF conversion failed: {str(e)}", exc_info=True)
            raise Exception(f"Conversion failed: {str(e)}")